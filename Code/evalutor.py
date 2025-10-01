import time
import json
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.formatting.rule import ColorScaleRule
import os
import pandas as pd
from pypdf import PdfReader
from google import genai
from sentence_transformers import SentenceTransformer
import faiss
import numpy as np


embedding_model = SentenceTransformer('all-MiniLM-L6-v2')
api_key = "set_your_api_key"
def chunk_pdf_by_words(pdf_path: str, chunk_size: int = 250, overlap: int = 50) -> pd.DataFrame:
    """
    Chunks a PDF into overlapping word-based segments with page tracking.

    Parameters:
        pdf_path (str): Path to the PDF file.
        chunk_size (int): Number of words per chunk.
        overlap (int): Number of overlapping words between chunks.

    Returns:
        pd.DataFrame with columns:
            - 'eval_chunk_id': str (e.g., 'eval_1')
            - 'eval_page_num': int (starting from 1)
            - 'eval_chunk_txt': str (cleaned text)
    """
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    reader = PdfReader(pdf_path)
    all_chunks = []
    chunk_id_counter = 1

    for page_num, page in enumerate(reader.pages, start=1):
        # Extract and clean text
        text = page.extract_text()
        if not text:
            continue
        # Basic cleaning: normalize whitespace
        text = re.sub(r'\s+', ' ', text).strip()
        words = text.split()

        if not words:
            continue

        # Create overlapping chunks
        start = 0
        while start < len(words):
            end = start + chunk_size
            chunk_words = words[start:end]
            chunk_text = " ".join(chunk_words)

            all_chunks.append({
                'eval_chunk_id': f"eval_{chunk_id_counter}",
                'eval_page_num': page_num,
                'eval_chunk_txt': chunk_text
            })
            chunk_id_counter += 1

            # Move start forward by (chunk_size - overlap), but avoid infinite loop
            next_start = start + (chunk_size - overlap)
            if next_start <= start:  # prevent zero/negative step
                next_start = start + 1
            start = next_start

    if not all_chunks:
        raise ValueError("No text could be extracted from the PDF.")

    df = pd.DataFrame(all_chunks)
    print(f"✅ Extracted {len(df)} chunks from {len(reader.pages)} pages.")
    return df
def load_rag_artifacts(
        index_name: str,
        base_dir: str = r"C:\Users\Thoma\Downloads\FireShot\Corpus"
) -> tuple[faiss.Index, pd.DataFrame, np.ndarray | None]:
    """
    Loads a FAISS index, its corresponding text chunks, and optionally embeddings.

    Parameters:
        index_name (str): Name of the index (e.g., 'england_nc' → looks for 'england_nc.faiss')
        base_dir (str): Base directory containing RAG_Data_Text, RAG_FAISS_Indices, RAG_Data_Embeddings

    Returns:
        tuple: (faiss_index, text_df, embeddings or None)
    """
    # Define paths
    faiss_dir = os.path.join(base_dir, "RAG_FAISS_Indices")
    text_dir = os.path.join(base_dir, "RAG_Data_Text")
    embed_dir = os.path.join(base_dir, "RAG_Data_Embeddings")

    faiss_path = os.path.join(faiss_dir, f"{index_name}.faiss")
    parquet_path = os.path.join(text_dir, f"{index_name}.parquet")
    csv_path = os.path.join(text_dir, f"{index_name}.csv")
    embed_path = os.path.join(embed_dir, f"{index_name}.npy")

    # Validate FAISS file exists
    if not os.path.exists(faiss_path):
        available = [f.replace('.faiss', '') for f in os.listdir(faiss_dir) if f.endswith('.faiss')]
        raise FileNotFoundError(
            f"FAISS index '{index_name}.faiss' not found in {faiss_dir}.\n"
            f"Available indexes: {available}"
        )

    # Load FAISS index
    print(f"Loading FAISS index: {faiss_path}")
    faiss_index = faiss.read_index(faiss_path)

    # Load text data (prefer Parquet, fallback to CSV)
    text_path = None
    if os.path.exists(parquet_path):
        text_path = parquet_path
        text_df = pd.read_parquet(parquet_path)
    elif os.path.exists(csv_path):
        text_path = csv_path
        text_df = pd.read_csv(csv_path)
    else:
        raise FileNotFoundError(f"No .parquet or .csv found for '{index_name}' in {text_dir}")

    print(f"Loaded text data from: {text_path} ({len(text_df)} chunks)")

    # Optional: Load embeddings
    embeddings = None
    if os.path.exists(embed_path):
        print(f"Loading embeddings from: {embed_path}")
        embeddings = np.load(embed_path)
        if embeddings.shape[0] != len(text_df):
            print("⚠️ Warning: Embedding count doesn't match text chunk count!")

    # Ensure required columns exist
    required_cols = {'chunk_id', 'chunk_txt', 'page_num'}
    if not required_cols.issubset(text_df.columns):
        print(f"⚠️ Warning: Text DataFrame missing columns. Found: {list(text_df.columns)}")

    return faiss_index, text_df, embeddings

def evaluate_chunk(eval_chunk_text: str, index, source_df: pd.DataFrame) -> dict:
    try:
        # 1. Embed using all-MiniLM (same as FAISS index)
        query_vector = embedding_model.encode([eval_chunk_text], convert_to_numpy=True).astype('float32')
        distances, indices = index.search(query_vector, k=3)

        retrieved_context_list = []
        for idx in indices[0]:
            if idx < len(source_df):
                chunk = source_df.iloc[idx]
                context = f"[Page {chunk['page_num']}] {chunk['chunk_txt']}"
                retrieved_context_list.append(context)
        retrieved_context = "\n".join(retrieved_context_list)

        # 2. Use Google's new SDK for generation
        client = genai.Client(api_key=api_key)  # reads GEMINI_API_KEY

        response = client.models.generate_content(
            model="gemini-2.5-flash-lite",
            contents=f"""
            You are an expert academic fact-checker. Evaluate the REPORT TEXT strictly against the provided SOURCE EXCERPTS.
            Respond ONLY with a valid JSON object with keys: "accuracy_score" (int 0-10), "helpful_comment" (str), "evidence_used" (str).
            Always return evidence used. If score is less than 5, frame it as suggested reading.

            SOURCE EXCERPTS:
            {retrieved_context}

            REPORT TEXT TO EVALUATE:
            "{eval_chunk_text}"
            """
        )
        time.sleep(5)
        raw_output = response.text.strip()
        if raw_output.startswith("```json"):
            raw_output = re.sub(r"^```json\s*|\s*```$", "", raw_output, flags=re.MULTILINE)
        result = json.loads(raw_output)

        return {
            'accuracy_score': int(result['accuracy_score']),
            'helpful_comment': result['helpful_comment'].strip(),
            'evidence_used': result['evidence_used'].strip()
        }

    except Exception as e:
        print(e)
        return {
            'accuracy_score': -1,
            'helpful_comment': f"Error: {str(e)}",
            'evidence_used': "N/A"
        }
def run_rag_evaluation(
        eval_df: pd.DataFrame,
        faiss_index,
        source_df: pd.DataFrame,
        chunk_text_col: str = 'eval_chunk_txt',
        chunk_id_col: str = 'eval_chunk_id',
        page_num_col: str = 'eval_page_num',
        delay: float = 0.5
) -> pd.DataFrame:
    """
    Runs RAG-based evaluation on each chunk in eval_df using a FAISS index and source documents.

    Parameters:
    - eval_df: DataFrame containing report chunks to evaluate.
    - faiss_index: Pre-built FAISS index over source document embeddings.
    - source_df: DataFrame of source chunks with 'chunk_txt', 'page_num', etc.
    - client: Google Generative AI client instance.
    - chunk_text_col, chunk_id_col, page_num_col: Column names in eval_df.
    - delay: Seconds to sleep between LLM calls (to respect rate limits).

    Returns:
    - DataFrame with columns: chunk_id, source_page, report_text, accuracy_score, helpful_comment, evidence_used
    """
    evaluation_results = []

    print("\n--- Starting RAG Evaluation Loop ---")
    for _, row in eval_df.iterrows():
        eval_id = row[chunk_id_col]
        eval_page = row[page_num_col]
        eval_text = row[chunk_text_col]

        print(f"Evaluating {eval_id} (Page {eval_page})...")

        try:
            result = evaluate_chunk(eval_text, faiss_index, source_df)
            evaluation_results.append({
                'chunk_id': eval_id,
                'source_page': eval_page,
                'report_text': eval_text,
                **result  # unpacks accuracy_score, helpful_comment, evidence_used
            })
        except Exception as e:
            print(f"⚠️ Error during evaluation for {eval_id}: {e}")
            evaluation_results.append({
                'chunk_id': eval_id,
                'source_page': eval_page,
                'report_text': eval_text,
                'accuracy_score': -1,
                'helpful_comment': f"Pipeline error: {str(e)}",
                'evidence_used': "N/A"
            })

        time.sleep(30)

    return pd.DataFrame(evaluation_results)

def save_evaluation_to_excel(df: pd.DataFrame, output_path: str = "rag_evaluation_scored_formatted.xlsx"):
    """
    Saves the evaluation DataFrame to an Excel file with:
    - Conditional color scale on 'accuracy_score' (red → green)
    - Auto-adjusted column widths
    - Wrapped text in long columns
    """
    # Ensure accuracy_score is numeric
    df = df.copy()
    df['accuracy_score'] = pd.to_numeric(df['accuracy_score'], errors='coerce').fillna(-1)

    # Save initial DataFrame to Excel
    df.to_excel(output_path, index=False, sheet_name="Evaluation")

    # Load workbook and worksheet
    wb = load_workbook(output_path)
    ws = wb["Evaluation"]

    # Identify column indices (1-based for openpyxl)
    col_names = [cell.value for cell in ws[1]]
    score_col_idx = col_names.index('accuracy_score') + 1 if 'accuracy_score' in col_names else None
    wrap_cols = ['report_text', 'helpful_comment', 'evidence_used']
    wrap_col_indices = [i + 1 for i, name in enumerate(col_names) if name in wrap_cols]

    # Apply conditional formatting (red-yellow-green) to accuracy_score
    if score_col_idx is not None:
        max_row = ws.max_row
        score_range = f"{ws.cell(2, score_col_idx).coordinate}:{ws.cell(max_row, score_col_idx).coordinate}"
        ws.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type='num', start_value=0, start_color='FF0000',   # Red
                mid_type='num', mid_value=5, mid_color='FFFF00',        # Yellow
                end_type='num', end_value=10, end_color='00FF00'        # Green
            )
        )

    # Apply text wrapping and alignment
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in wrap_col_indices:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            # Optional: bold header
            if row[0].row == 1:
                cell.font = Font(bold=True)

    # Auto-adjust column widths (with reasonable max)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get column letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)  # Cap at 60 chars
        ws.column_dimensions[column].width = adjusted_width

    # Save final formatted workbook
    wb.save(output_path)
    print(f"✅ Formatted Excel saved to: {os.path.abspath(output_path)}")
    full_path = os.path.abspath(output_path)
    return full_path

def evaluate_report_against_rag_index(
        base_corpus_dir: str = r"C:\Users\Thoma\Downloads\FireShot\Corpus",
        eval_chunk_size: int = 250,
        eval_overlap: int = 50,
        output_excel_name: str = "evaluation_output.xlsx"
):
    """
    Interactive function to evaluate a report PDF against a chosen RAG FAISS index.

    Parameters:
        base_corpus_dir (str): Base directory containing RAG_FAISS_Indices, RAG_Data_Text, etc.
        eval_chunk_size (int): Words per evaluation chunk.
        eval_overlap (int): Overlap between chunks (in words).
        output_excel_name (str): Name of the output Excel file.

    Returns:
        str: Full path to the saved Excel file.
    """
    # --- Step 1: List available FAISS indexes ---
    faiss_dir = os.path.join(base_corpus_dir, "RAG_FAISS_Indices")
    if not os.path.exists(faiss_dir):
        raise FileNotFoundError(f"FAISS directory not found: {faiss_dir}")

    available_indexes = [f.replace('.faiss', '') for f in os.listdir(faiss_dir) if f.endswith('.faiss')]
    if not available_indexes:
        raise ValueError(f"No .faiss files found in {faiss_dir}")

    print("Available RAG indexes:")
    for i, name in enumerate(available_indexes, 1):
        print(f"  {i}. {name}")

    # --- Step 2: Get user input ---
    while True:
        try:
            index_choice = input("\nEnter the name of the FAISS index to use (e.g., England): ").strip()
            if index_choice in available_indexes:
                break
            else:
                print("❌ Invalid index name. Please choose from the list above.")
        except KeyboardInterrupt:
            print("\nOperation cancelled.")
            return None

    report_path = input("Enter the full path to the report PDF: ").strip()
    if not os.path.exists(report_path):
        raise FileNotFoundError(f"Report PDF not found: {report_path}")

    # --- Step 3: Load RAG artifacts ---
    faiss_index, source_df, _ = load_rag_artifacts(
        index_name=index_choice,
        base_dir=base_corpus_dir
    )

    # --- Step 4: Chunk the report PDF ---
    eval_df = chunk_pdf_by_words(
        pdf_path=report_path,
        chunk_size=eval_chunk_size,
        overlap=eval_overlap
    )

    # --- Step 5: Run RAG evaluation ---
    final_df = run_rag_evaluation(
        eval_df=eval_df,
        faiss_index=faiss_index,
        source_df=source_df,
        delay=0.5
    )

    # --- Step 6: Add average row ---
    final_df['accuracy_score'] = pd.to_numeric(final_df['accuracy_score'], errors='coerce')
    valid_scores = final_df['accuracy_score'].replace(-1, pd.NA).dropna()
    average_score = valid_scores.mean()

    summary_row = {
        'chunk_id': 'AVERAGE',
        'source_page': '',
        'report_text': 'Overall Average Accuracy Score',
        'accuracy_score': round(average_score, 2) if not pd.isna(average_score) else 0,
        'helpful_comment': f"Average of {len(valid_scores)} valid evaluations.",
        'evidence_used': ''
    }

    final_df_with_avg = pd.concat([final_df, pd.DataFrame([summary_row])], ignore_index=True)

    # --- Step 7: Save to formatted Excel ---
    full_path = save_evaluation_to_excel(final_df_with_avg, output_excel_name)
    print(f"✅ Evaluation complete. Saved to: {full_path}")
    return full_path

def evaluate_custom_report(pdf_path: str, index_name: str, output_name: str):
    """
    Interactive function to evaluate any report PDF against any available FAISS RAG index.
    Prompts user for:
      - Path to the report PDF (e.g., Curriculum for Wales)
      - Name of the FAISS index (e.g., 'Wales')
    """
    base_dir = r"C:\Users\Thoma\Downloads\FireShot\Corpus"
    # --- List available indexes ---
    faiss_dir = os.path.join(base_dir, "RAG_FAISS_Indices")
    available = [f.replace('.faiss', '') for f in os.listdir(faiss_dir) if f.endswith('.faiss')]
    print("Available RAG indexes:", available)

    # --- Get inputs ---

    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    # --- Load RAG artifacts ---
    faiss_index, source_df, _ = load_rag_artifacts(index_name=index_name, base_dir=base_dir)

    # --- Chunk report ---
    eval_df = chunk_pdf_by_words(pdf_path, chunk_size=45, overlap=10)

    # --- Evaluate ---
    final_df = run_rag_evaluation(eval_df, faiss_index, source_df, delay=0.5)

    # --- Add average ---
    final_df['accuracy_score'] = pd.to_numeric(final_df['accuracy_score'], errors='coerce')
    valid_scores = final_df['accuracy_score'].replace(-1, pd.NA).dropna()
    avg_score = round(valid_scores.mean(), 2) if not pd.isna(valid_scores.mean()) else 0

    summary = pd.DataFrame([{
        'chunk_id': 'AVERAGE',
        'source_page': '',
        'report_text': 'Overall Average Accuracy Score',
        'accuracy_score': avg_score,
        'helpful_comment': f"Average of {len(valid_scores)} valid evaluations.",
        'evidence_used': ''
    }])

    final_with_avg = pd.concat([final_df, summary], ignore_index=True)

    # --- Save ---
    full_path = save_evaluation_to_excel(final_with_avg, output_name)
    print(f"✅ Done! Saved to: {full_path}")
    return full_path

