import time
print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] 🚀 Starting application imports...", flush=True)

# --- 1. Core Python Standard Library ---
print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] ➤ Importing core modules: os, json, re, random, tempfile, shutil, subprocess, sys...", flush=True)
import os
import json
import re
import random
import tempfile
import shutil
import subprocess
import sys

# --- 2. Numerical & Data Processing ---
print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] ➤ Importing numerical & data libraries: numpy, pandas...", flush=True)
import numpy as np
import pandas as pd

# --- 3. File & Path Handling ---
print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] ➤ Importing pathlib for path operations...", flush=True)
import pathlib
from pathlib import Path

# --- 4. Document Processing ---
print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] ➤ Importing PDF and Excel libraries...", flush=True)
from pypdf import PdfReader
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.formatting.rule import ColorScaleRule

# --- 5. Vector Search ---
print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] ➤ Importing FAISS for vector similarity search...", flush=True)
import faiss

# --- 6. NLP: NLTK (pre-cached, no download at runtime) ---
print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] ➤ Setting NLTK_DATA and importing nltk...", flush=True)
os.environ.setdefault("NLTK_DATA", "/app/nltk_data")
import nltk  # ✅ Fixed typo: was "nlkt"

# --- 7. AI / Embeddings (DELAYED IMPORT - load only when needed) ---
# We will import SentenceTransformer later in a function to avoid startup hang

# --- 8. Google GenAI ---
print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] ➤ Importing Google GenAI...", flush=True)
try:
    from google import genai
    from google.genai import types
except ImportError as e:
    print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] ⚠️ Google GenAI not available: {e}", flush=True)
    genai = None

# --- 9. Gradio (UI Layer - load last) ---
print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] ➤ Importing Gradio for UI...", flush=True)
import gradio as gr

print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] ✅ All imports completed successfully.", flush=True)


def cleanup_cache():
    # Use HF_HOME if set, fallback to default
    hf_home = os.environ.get("HF_HOME", str(Path.home() / ".cache" / "huggingface"))
    cache_dir = Path(hf_home)
    
    if not cache_dir.exists():
        print(f"Cache dir {cache_dir} does not exist. Skipping cleanup.", flush=True)
        return

    try:
        print(f"Checking size of cache dir: {cache_dir}", flush=True)
        
        # Add a safeguard: skip if too many files (avoid hang)
        file_count = 0
        total_size = 0
        for f in cache_dir.rglob('*'):
            if f.is_file():
                file_count += 1
                total_size += f.stat().st_size
                # Optional: early break if already over threshold
                if total_size > 10 * 1024**3:
                    break
                # Optional: prevent excessive scanning
                if file_count > 10000:
                    print("Too many files (>10k), skipping detailed size check.", flush=True)
                    return

        print(f"Cache size: {total_size / (1024**3):.2f} GB ({file_count} files)", flush=True)
        
        if total_size > 10 * 1024**3:
            print(f"Cache exceeds 10 GB. Deleting {cache_dir}...", flush=True)
            shutil.rmtree(cache_dir)
            print("✅ Cache deleted.", flush=True)
        else:
            print("Cache within limit. No cleanup needed.", flush=True)

    except Exception as e:
        print(f"⚠️ Cache cleanup failed: {e}", flush=True)
cleanup_cache()
# Define the NLTK data path where it was installed in the Dockerfile
# We point to the exact location: /app/nltk_data
#nltk_data_dir = "/app/nltk_data" 
#nltk.data.path.append(nltk_data_dir)

# NOTE: The nltk.download() call is REMOVED, as the data is already pre-installed.



print("  - Importing SentenceTransformer class...")
from sentence_transformers import SentenceTransformer

print("  - Calling SentenceTransformer('all-MiniLM-L6-v2') NOW...")
print("    ⏳ This may take 10-30 seconds on first load. DO NOT KILL.")

embedding_model = SentenceTransformer('all-MiniLM-L6-v2')

list_of_models = ["gemini-2.5-flash-lite"]



def convert_docx_to_pdf(docx_path, output_dir):
    """Converts a DOCX file to PDF using LibreOffice in headless mode (HF Spaces compatible)."""
    docx_path = Path(docx_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    pdf_path = output_dir / (docx_path.stem + ".pdf")
    print(f"\n--- Starting DOCX to PDF Conversion for: {docx_path.name} ---")

    if pdf_path.exists():
        print(f"   --> Found converted PDF: {pdf_path.name}. Skipping conversion.")
        return pdf_path

    # Enhanced LibreOffice command for container safety
    command = [
        'libreoffice',
        '--headless',
        '--invisible',
        '--norestore',
        '--nologo',
        '--nodefault',
        '-env:UserInstallation=file:///tmp',
        '--convert-to', 'pdf',
        '--outdir', str(output_dir),
        str(docx_path)
    ]

    try:
        print(f"   --> Running command: {' '.join(command)}")
        
        # Add timeout (e.g., 60 seconds) to avoid hanging
        result = subprocess.run(
            command,
            check=True,
            capture_output=True,
            text=True,
            timeout=60
        )

        # Wait briefly and check
        time.sleep(2)
        
        if pdf_path.exists():
            print(f"   --> 🎉 Successfully converted to PDF: {pdf_path.name}")
            return pdf_path
        else:
            print(f"   --> ❌ Output PDF not found at {pdf_path}")
            print(f"   --> STDOUT: {result.stdout}")
            print(f"   --> STDERR: {result.stderr}")
            return None

    except subprocess.TimeoutExpired:
        print("   --> ❌ Conversion timed out (60s).")
        return None
    except FileNotFoundError:
        print("   --> ❌ LibreOffice not found. Not installed in environment.")
        return None
    except subprocess.CalledProcessError as e:
        print(f"   --> ❌ LibreOffice failed with code {e.returncode}")
        print(f"   --> STDERR: {e.stderr}")
        return None
    except Exception as e:
        print(f"   --> ❌ Unexpected error: {e}")
        return None
        

def clean_text(t):
    # Removes leading/trailing whitespace and normalizes newlines
    return re.sub(r'^\s+|\s+$', '', t).replace('\n', ' ')
def chunk_pdf_by_words(pdf_path: str, chunk_size: int = 250, overlap: int = 50) -> pd.DataFrame:
    """
    Chunks a PDF using sentence-aware (semantic fragment) aggregation.
    The logic ensures every chunk is built from complete sentences/phrases
    and meets a minimum size threshold.
    Parameters:
        pdf_path (str): Path to the PDF file.
        chunk_size (int): **Minimum** word count for a chunk (default 250).
        overlap (int): Ignored in this sentence-aware aggregation logic.
    Returns:
        pd.DataFrame with columns:
            - 'eval_chunk_id': str (e.g., 'eval_1')
            - 'eval_page_num': int (starting from 1)
            - 'eval_chunk_txt': str (cleaned text)
    """
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF not found: {pdf_path}")
    # --- 1. Define Semantic Splitters (Connectives and Punctuation) ---
    connectives = r'\b(?:however|therefore|moreover|but|although|yet|because|so)\b'
    # Split on major punctuation and strong connectives to get sentence-like fragments
    split_regex = r'[;,\.\?!:\-\–\n]+' + '|' + connectives
    reader = PdfReader(pdf_path)
    all_chunks = []
    chunk_id_counter = 1
    # The minimum word count threshold (default 250, or your specific 9 words)
    MIN_CHUNK_WORDS = chunk_size
    # Variable to hold text that spills over from page to page
    page_spillover_buffer = ""
    # --- 2. Iterate Pages, Fragment, and Aggregate ---
    for page_num, page in enumerate(reader.pages, start=1):
        text = page.extract_text()
        if not text:
            continue
        # Prepend spillover from the previous page
        full_text = page_spillover_buffer + " " + text
        page_spillover_buffer = ""  # Clear the spillover buffer
        # Split the text into semantic fragments
        fragments = re.split(split_regex, full_text, flags=re.IGNORECASE)
        # The last fragment might be incomplete (split by page end or final regex)
        # We hold the LAST fragment back as potential spillover for the next page,
        # UNLESS it's the very last page.
        fragments_to_process = fragments[:-1]
        last_fragment = fragments[-1]
        # On the last page, process the final fragment as well
        if page_num == len(reader.pages):
            fragments_to_process.append(last_fragment)
        else:
            # If not the last page, the final fragment spills over
            page_spillover_buffer = clean_text(last_fragment)
        # --- 3. Aggregate Fragments into Chunks (Min Word Count Logic) ---
        aggregation_buffer = ""
        for fragment in fragments_to_process:
            cleaned_fragment = clean_text(fragment)
            if not cleaned_fragment:
                continue
            # Add fragment to the buffer
            aggregation_buffer += cleaned_fragment + " "
            # Check the buffer size
            words = nltk.word_tokenize(aggregation_buffer)
            length = len(words)
            # If buffer meets the minimum size, finalize it as a chunk
            if length >= MIN_CHUNK_WORDS:
                all_chunks.append({
                    'eval_chunk_id': f"eval_{chunk_id_counter}",
                    # Assign the page number where the chunk started/finished
                    'eval_page_num': page_num,
                    'eval_chunk_txt': aggregation_buffer.strip()
                })
                chunk_id_counter += 1
                aggregation_buffer = ""  # Reset the buffer
    # --- 4. Final Residue Spill-Over Logic (Fixes the Critical Error) ---
    if aggregation_buffer.strip():
        residue_text = aggregation_buffer.strip()
        residue_words = nltk.word_tokenize(residue_text)
        # Check if residue is too small (e.g., less than 50% of MIN_CHUNK_WORDS)
        # AND if there is a previous chunk to merge with
        MIN_RESIDUE_RATIO = 0.5
        if len(residue_words) < (MIN_CHUNK_WORDS * MIN_RESIDUE_RATIO) and len(all_chunks) > 0:
            # Spill-over: Append the residue to the LAST valid chunk
            last_chunk = all_chunks[-1]
            last_chunk['eval_chunk_txt'] += " " + residue_text
        else:
            # The residue is large enough, or it's the only content; keep it as a final chunk
            all_chunks.append({
                'eval_chunk_id': f"eval_{chunk_id_counter}",
                'eval_page_num': len(reader.pages),  # Assign to the last page processed
                'eval_chunk_txt': residue_text
            })
    if not all_chunks:
        raise ValueError("No text could be extracted or chunked from the PDF.")
    df = pd.DataFrame(all_chunks)
    print(f"✅ Extracted {len(df)} sentence-aware chunks from {len(reader.pages)} pages.")
    return df



def load_rag_artifacts(
        index_name: str,
        base_dir: str = r"Data"
) -> tuple[faiss.Index, pd.DataFrame, np.ndarray | None]:
    """
    Loads a FAISS index, its corresponding text chunks, and optionally embeddings.

    Parameters:
        index_name (str): Name of the index (e.g., 'england_nc' → looks for 'england_nc.faiss')
        base_dir (str): Base directory containing RAG_Data_Text, RAG_FAISS_Indices, RAG_Data_Embeddings

    Returns:
        tuple: (faiss_index, text_df, embeddings or None)
    """
    # Define paths
    faiss_dir = os.path.join(base_dir, "RAG_FAISS_Indices")
    text_dir = os.path.join(base_dir, "RAG_Data_Text")
    embed_dir = os.path.join(base_dir, "RAG_Data_Embeddings")

    faiss_path = os.path.join(faiss_dir, f"{index_name}.faiss")
    parquet_path = os.path.join(text_dir, f"{index_name}.parquet")
    csv_path = os.path.join(text_dir, f"{index_name}.csv")
    embed_path = os.path.join(embed_dir, f"{index_name}.npy")

    # Validate FAISS file exists
    if not os.path.exists(faiss_path):
        available = [f.replace('.faiss', '') for f in os.listdir(faiss_dir) if f.endswith('.faiss')]
        raise FileNotFoundError(
            f"FAISS index '{index_name}.faiss' not found in {faiss_dir}.\n"
            f"Available indexes: {available}"
        )

    # Load FAISS index
    print(f"Loading FAISS index: {faiss_path}")
    faiss_index = faiss.read_index(faiss_path)

    # Load text data (prefer Parquet, fallback to CSV)
    text_path = None
    if os.path.exists(parquet_path):
        text_path = parquet_path
        text_df = pd.read_parquet(parquet_path)
    elif os.path.exists(csv_path):
        text_path = csv_path
        text_df = pd.read_csv(csv_path)
    else:
        raise FileNotFoundError(f"No .parquet or .csv found for '{index_name}' in {text_dir}")

    print(f"Loaded text data from: {text_path} ({len(text_df)} chunks)")

    # Optional: Load embeddings
    embeddings = None
    if os.path.exists(embed_path):
        print(f"Loading embeddings from: {embed_path}")
        embeddings = np.load(embed_path)
        if embeddings.shape[0] != len(text_df):
            print("⚠️ Warning: Embedding count doesn't match text chunk count!")

    # Ensure required columns exist
    required_cols = {'chunk_id', 'chunk_txt', 'page_num'}
    if not required_cols.issubset(text_df.columns):
        print(f"⚠️ Warning: Text DataFrame missing columns. Found: {list(text_df.columns)}")

    return faiss_index, text_df, embeddings
    

def evaluate_chunk(eval_chunk_text: str, index, source_df: pd.DataFrame) -> dict:
    """
    Evaluates a report chunk against context retrieved from a FAISS index
    using the Gemini API. Includes robust retry logic and outputs a detailed 
    debug report to the 'helpful_comment' column upon failure.
    """
    MAX_RETRIES = 10
    
    # Initialize variables to capture state across attempts for the debug report
    last_error_message = "No attempt made." 
    retrieved_context = "Not yet retrieved."
    raw_output = "No API response received."
    
    for attempt in range(MAX_RETRIES):
        try:
            if attempt >= 1:
                api_key = os.environ["GEMINI_API_PAID"]
            else:
                api_key = os.environ["GEMINI_API_KEY"]
                time.sleep(4.2)
                
            # 1. Embed and Search FAISS Index
            query_vector = embedding_model.encode([eval_chunk_text], convert_to_numpy=True).astype('float32')
            distances, indices = index.search(query_vector, k=7)
            retrieved_context_list = []
            for idx in indices[0]:
                if idx < len(source_df):
                    chunk = source_df.iloc[idx]
                    context = f"[Page {chunk['page_num']}] {chunk['chunk_txt']}"
                    retrieved_context_list.append(context)
            
            retrieved_context = "\n".join(retrieved_context_list)
            
            # 2. Configure and Call Gemini API
            client = genai.Client(api_key=api_key)
            
            response = client.models.generate_content(
                config=types.GenerateContentConfig(temperature=0.4),
                model=random.choice(list_of_models),
                contents=f"""
                You are marking reports based on accuracy. 
                Evaluate the REPORT TEXT strictly against the provided SOURCE EXCERPTS.
                SOURCE EXCERPTS are the original documents the REPORT TEXT is based on.
                REPORT TEXT will be chunked, you will recieve discrete chunks.
                Do not penalise what may be a truncated sentence.
                REPORT TEXT does not have to contain all the facts/information from SOURCE EXCERPTS.
                Accept paraphrasing, where the meaning is retained.
                Penalise false claims, incorrect dates and misinterpretation.
                Respond ONLY with a valid JSON object with keys: "accuracy_score" (int 0-10), "helpful_comment" (str), "evidence_used" (str).
                Always return evidence used. This will include direction quotations and page numbers.
                SOURCE EXCERPTS:
                {retrieved_context}
                REPORT TEXT TO EVALUATE:
                "{eval_chunk_text}"
                """
            )
            
            
            # 3. Process Response and JSON Parse
            raw_output = response.text.strip()
            
            if raw_output.startswith("```json"):
                raw_output = re.sub(r"^```json\s*|\s*```$", "", raw_output, flags=re.MULTILINE)
                
            result = json.loads(raw_output)
            
            return {
                'accuracy_score': int(result['accuracy_score']),
                'helpful_comment': result['helpful_comment'],
                'explanations': result['evidence_used'],
                'references': retrieved_context
            }
            
        except Exception as e:
            last_error_message = f"Type: {type(e).__name__} | Message: {str(e)}"
            print(f'Attempt {attempt + 1}/{MAX_RETRIES} failed with error: {last_error_message}')
            print(f"RAW API OUTPUT captured on failure: {raw_output}")

            
    # 5. FINAL FAILURE
    debug_report = (
        f"*** PIPELINE ERROR: ALL {MAX_RETRIES} ATTEMPTS FAILED ***\n"
        f"LAST ERROR: {last_error_message}\n\n"
        f"--- RAW API RESPONSE (Check for Malformed JSON) ---\n"
        f"{raw_output}\n\n"
        f"--- FAISS CONTEXT USED (First 300 chars) ---\n"
        f"{retrieved_context[:300]}..."
    )
    
    return {
        'accuracy_score': -1,
        'helpful_comment': debug_report,
        'explanations': "N/A",
        'references': retrieved_context 
    }    

def run_rag_evaluation(
        eval_df: pd.DataFrame,
        faiss_index,
        source_df: pd.DataFrame,
        chunk_text_col: str = 'eval_chunk_txt',
        chunk_id_col: str = 'eval_chunk_id',
        page_num_col: str = 'eval_page_num',
        delay: float = 0.5
) -> pd.DataFrame:
    """
    Runs RAG-based evaluation on each chunk in eval_df using a FAISS index and source documents.

    Parameters:
    - eval_df: DataFrame containing report chunks to evaluate.
    - faiss_index: Pre-built FAISS index over source document embeddings.
    - source_df: DataFrame of source chunks with 'chunk_txt', 'page_num', etc.
    - client: Google Generative AI client instance.
    - chunk_text_col, chunk_id_col, page_num_col: Column names in eval_df.
    - delay: Seconds to sleep between LLM calls (to respect rate limits).

    Returns:
    - DataFrame with columns: chunk_id, source_page, report_text, accuracy_score,
      helpful_comment, evidence_used, faiss_retrieved_chunks 🚀
    """
    evaluation_results = []

    print("\n--- Starting RAG Evaluation Loop ---")
    for _, row in eval_df.iterrows():
        eval_id = row[chunk_id_col]
        eval_page = row[page_num_col]
        eval_text = row[chunk_text_col]

        print(f"Evaluating {eval_id} (Page {eval_page})...")

        try:
            # Assumes embedding_model is available here or passed to evaluate_chunk
            result = evaluate_chunk(eval_text, faiss_index, source_df)
            evaluation_results.append({
                'chunk_id': eval_id,
                'source_page': eval_page,
                'report_text': eval_text,
                # **result now correctly unpacks 'faiss_retrieved_chunks' along with
                # accuracy_score, helpful_comment, and evidence_used.
                **result
            })
        except Exception as e:
            print(f"⚠️ Error during evaluation for {eval_id}: {e}")
            evaluation_results.append({
                'chunk_id': eval_id,
                'source_page': eval_page,
                'report_text': eval_text,
                'accuracy_score': -1,
                'helpful_comment': f"Pipeline error: {str(e)}",
                'explanations': "N/A",
                'references': "N/A" # Added error handling for the new column
            })
    return pd.DataFrame(evaluation_results)


def save_evaluation_to_excel(df: pd.DataFrame, output_path: str = "rag_evaluation_scored_formatted.xlsx"):
    """
    Saves the evaluation DataFrame to an Excel file with:
    - Conditional color scale on 'accuracy_score' (red → green)
    - Auto-adjusted column widths. 'references' column is targeted to be twice as wide
      as the standard columns, both in auto-fit calculation and max width cap.
    - Wrapped text in long columns
    - Compact but readable row heights
    """
    df = df.copy()
    # Convert score to numeric, coercing errors and filling missing values with -1
    df['accuracy_score'] = pd.to_numeric(df['accuracy_score'], errors='coerce').fillna(-1)

    # Use a safe temp path from the start — critical for Gradio/Spaces
    temp_dir = tempfile.gettempdir()
    safe_output_path = os.path.join(temp_dir, "rag_evaluation_scored_formatted.xlsx")
    print(f"📝 Saving initial Excel to: {safe_output_path}")

    # Save raw Excel first (no formatting)
    df.to_excel(safe_output_path, index=False, sheet_name="Evaluation")

    # Load for formatting
    wb = load_workbook(safe_output_path)
    ws = wb["Evaluation"]
    print("✅ Workbook loaded")

    # --- Column indexing ---
    col_names = [cell.value for cell in ws[1]]
    print(f".Columns found: {col_names}")

    score_col_idx = col_names.index('accuracy_score') + 1 if 'accuracy_score' in col_names else None
    references_col_idx = col_names.index('references') + 1 if 'references' in col_names else None
    wrap_cols = ['report_text', 'helpful_comment', 'explanations', 'references']
    wrap_col_indices = [i + 1 for i, name in enumerate(col_names) if name in wrap_cols]

    print(f"🎯 score_col_idx: {score_col_idx}, references_col_idx: {references_col_idx}")
    print(f"📄 Wrap columns (1-based): {wrap_col_indices}")

    col_width_map = {}
    MAX_STANDARD_WIDTH = 50
    MAX_REFERENCES_WIDTH = 100

    # --- 1. Conditional formatting ---
    if score_col_idx is not None and ws.max_row >= 2:
        score_range = f"{ws.cell(2, score_col_idx).coordinate}:{ws.cell(ws.max_row, score_col_idx).coordinate}"
        print(f"🎨 Applying color scale to: {score_range}")
        ws.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type='num', start_value=0, start_color='FF0000',
                mid_type='num', mid_value=5, mid_color='FFFF00',
                end_type='num', end_value=10, end_color='00FF00'
            )
        )
    else:
        print("⚠️  Skipping conditional formatting (missing column or too few rows)")

    # --- 2. Text wrapping ---
    wrap_count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in wrap_col_indices:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                wrap_count += 1
    print(f"↩️  Applied text wrapping to {wrap_count} cells")

    # --- 3. Bold header ---
    for cell in ws[1]:
        cell.font = Font(bold=True)
    print("✨ Header bolded")

    # --- 4. Column widths ---
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        col_idx = col[0].column

        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                line_len = max(len(line) for line in lines) if lines else 0
                max_length = max(max_length, line_len)

        base_width = max_length + 4
        if col_idx == references_col_idx:
            width = min(base_width * 2, MAX_REFERENCES_WIDTH)
        else:
            width = min(base_width, MAX_STANDARD_WIDTH)

        ws.column_dimensions[col_letter].width = width
        col_width_map[col_idx] = width
        print(f"↔️  Column {col_letter} width = {width}")

    # --- 5. Row heights ---
    DEFAULT_ROW_HEIGHT = 15
    MIN_ROW_HEIGHT = 20
    MAX_ROW_HEIGHT = 200
    CHARS_PER_UNIT = 1.3

    for row_idx in range(2, ws.max_row + 1):
        max_lines = 1
        for col_idx in wrap_col_indices:
            cell = ws.cell(row=row_idx, column=col_idx)
            content = str(cell.value) if cell.value else ""
            if not content.strip():
                continue

            col_w = col_width_map.get(col_idx, 15)
            lines_needed = 0
            for seg in content.split('\n'):
                if seg == "":
                    lines_needed += 1
                else:
                    chars_per_line = max(1, int(col_w * CHARS_PER_UNIT))
                    lines_needed += (len(seg) + chars_per_line - 1) // chars_per_line
            max_lines = max(max_lines, lines_needed)

        height = max(MIN_ROW_HEIGHT, min(max_lines * DEFAULT_ROW_HEIGHT, MAX_ROW_HEIGHT))
        ws.row_dimensions[row_idx].height = height
        # Uncomment below if you want row height logs (can be noisy)
        # print(f"↕️  Row {row_idx} height = {height}")

    # --- Save final file ---
    wb.save(safe_output_path)
    print(f"✅ Final formatted Excel saved to: {safe_output_path}")

    # RETURN THE ACTUAL SAVED PATH — this is critical for Gradio!
    return safe_output_path
def evaluate_report_against_rag_index(
    report_pdf_path: str,
    eval_chunk_size: int = 250,
    eval_overlap: int = 50,
    output_excel_name: str = "rag_evaluation_scored_formatted.xlsx"
) -> str:
    """
    Evaluates a report PDF against a RAG index built **from the same PDF** (self-evaluation).
    Designed for Hugging Face Spaces: no interactive input, no local paths assumed.
    Parameters:
        report_pdf_path (str): Path to uploaded PDF (from Gradio).
        eval_chunk_size (int): Min words per chunk.
        eval_overlap (int): Ignored in current chunking logic.
        output_excel_name (str): Output filename (will be saved to temp dir).
    Returns:
        str: Path to saved evaluation Excel file (in temp directory).
    """
    # --- Step 1: Chunk the report PDF ---
    eval_df = chunk_pdf_by_words(
        pdf_path=report_pdf_path,
        chunk_size=eval_chunk_size,
        overlap=eval_overlap
    )

    # --- Step 2: Use the SAME chunks as both query and source (self-RAG) ---
    source_df = eval_df.rename(columns={
        'eval_chunk_id': 'chunk_id',
        'eval_page_num': 'page_num',
        'eval_chunk_txt': 'chunk_txt'
    }).copy()

    # --- Step 3: Embed source chunks ---
    print("Embedding source chunks...")
    source_texts = source_df['chunk_txt'].tolist()
    source_embeddings = embedding_model.encode(source_texts, show_progress_bar=False)

    # Build FAISS index (flat L2)
    dim = source_embeddings.shape[1]
    faiss_index = faiss.IndexFlatL2(dim)
    faiss_index.add(source_embeddings.astype(np.float32))

    # --- Step 4: Run RAG evaluation ---
    final_df = run_rag_evaluation(
        eval_df=eval_df,
        faiss_index=faiss_index,
        source_df=source_df,
        delay=0.0  # No delay needed for local model
    )

    # --- Step 5: Add average row ---
    final_df['accuracy_score'] = pd.to_numeric(final_df['accuracy_score'], errors='coerce')
    valid_scores = final_df['accuracy_score'].replace(-1, pd.NA).dropna()
    average_score = valid_scores.mean() if not valid_scores.empty else 0

    summary_row = {
        'chunk_id': 'AVERAGE',
        'source_page': '',
        'report_text': 'Overall Average Accuracy Score',
        'accuracy_score': round(average_score, 2),
        'helpful_comment': f"Average of {len(valid_scores)} valid evaluations.",
        'explanations': '',
        'references': ''
    }

    final_df_with_avg = pd.concat([final_df, pd.DataFrame([summary_row])], ignore_index=True)

    # --- Step 6: Save to formatted Excel ---
    full_path = save_evaluation_to_excel(final_df_with_avg, output_excel_name)
    return full_path

def evaluate_custom_report(
    word_input,  # gr.File object from Gradio
    index_name: str,
    output_name: str,
    base_dir: str = "Data",
    google_api_key: str = None
) -> str:
    """
    Evaluates a report DOCX against a pre-built FAISS RAG index.
    Designed for Gradio: safe paths, Gemini support, no interactivity.

    Parameters:
        word_input (gr.File): Uploaded DOCX file object.
        index_name (str): Name of FAISS index (e.g., 'Wales').
        output_name (str): Output filename (e.g., 'wales_eval.xlsx').
        base_dir (str): Base directory for RAG artifacts (default: 'Data').
        google_api_key (str): Google API key for Gemini.

    Returns:
        str: Path to saved evaluation Excel file (in temp directory).
    """
    start_time = time.time()

    # --- Configure Gemini API key ---
    if google_api_key:
        os.environ["GEMINI_API_KEY"] = google_api_key
    elif not os.environ.get("GEMINI_API_KEY"):
        raise ValueError("Google API key not provided and not set in environment.")

    # --- Convert DOCX to PDF ---
    if not word_input:
        raise ValueError("No DOCX file uploaded.")

    temp_dir = tempfile.mkdtemp()
    uploaded_docx_path = Path(word_input.name)
    pdf_output_path = convert_docx_to_pdf(uploaded_docx_path, Path(temp_dir))
    if not pdf_output_path:
        raise ValueError("Failed to convert DOCX to PDF.")

    # --- Load RAG artifacts ---
    faiss_index, source_df, _ = load_rag_artifacts(index_name=index_name, base_dir=base_dir)

    # --- Chunk report ---
    eval_df = chunk_pdf_by_words(str(pdf_output_path), chunk_size=45, overlap=10)

    # --- Evaluate using Gemini (via global embedding_model and configured genai) ---
    final_df = run_rag_evaluation(
        eval_df=eval_df,
        faiss_index=faiss_index,
        source_df=source_df,
        delay=0.5
    )

    # --- Add average row ---
    final_df['accuracy_score'] = pd.to_numeric(final_df['accuracy_score'], errors='coerce')
    valid_scores = final_df['accuracy_score'].replace(-1, pd.NA).dropna()
    avg_score = round(valid_scores.mean(), 2) if not valid_scores.empty else 0

    # ⚠️ CRITICAL FIX: Use 'explanations' and 'references' to match evaluate_chunk output
    summary = pd.DataFrame([{
        'chunk_id': 'AVERAGE',
        'source_page': '',
        'report_text': 'Overall Average Accuracy Score',
        'accuracy_score': avg_score,
        'helpful_comment': f"Average of {len(valid_scores)} valid evaluations.",
        'explanations': '',          # ← was 'evidence_used' (wrong key)
        'references': '[]'           # ← keep consistent
    }])

    final_with_avg = pd.concat([final_df, summary], ignore_index=True)

    # --- Save to temporary location ---
    full_path = save_evaluation_to_excel(final_with_avg, output_name)

    # --- Timing ---
    end_time = time.time()
    run_time = (end_time - start_time) / 60
    print(f"✅ Done! Saved to: {full_path}")
    print(f"⏱️  Runtime: {run_time:.2f} minutes")

    return full_path

# Import your functions (make sure they're in this file or imported)
# Assuming all functions (chunk_pdf_by_words, load_rag_artifacts, etc.) are defined above or in this script

# -----------------------------
# Helper: List available indexes
# -----------------------------
def get_available_indexes(base_dir: str = "Data") -> list:
    faiss_dir = Path(base_dir) / "RAG_FAISS_Indices"
    if not faiss_dir.exists():
        return []
    return [f.stem for f in faiss_dir.glob("*.faiss")]

# -----------------------------
# Main Gradio function
# -----------------------------
def run_evaluation(word_file, index_name):
    if not word_file:
        raise gr.Error("Please upload a DOCX report.")
    if not index_name:
        raise gr.Error("Please select a RAG index.")

    google_api_key = os.environ.get('GEMINI_API_KEY')
    if not google_api_key:
        raise gr.Error("Google API key is required. Set it as a Hugging Face Secret named 'GEMINI_API_KEY'.")

    try:
        result_path = evaluate_custom_report(
            word_input=word_file,          # ← DOCX file object from Gradio
            index_name=index_name,
            output_name="rag_evaluation_results.xlsx",
            base_dir="Data",
            google_api_key=google_api_key
        )
        return result_path
    except Exception as e:
        raise gr.Error(f"Evaluation failed: {str(e)}")


# -----------------------------
# Build Gradio Interface
# -----------------------------
available_indexes = get_available_indexes()

# If no indexes found, show warning
if not available_indexes:
    available_indexes = ["No indexes found — check data/RAG_FAISS_Indices"]

with gr.Blocks(title="RAG Report Evaluator") as demo:
    gr.Markdown("# 📄 RAG Report Evaluator")
    gr.Markdown("Upload a policy or curriculum PDF and evaluate its alignment with a reference knowledge base using **Gemini-powered RAG**.")

    with gr.Row():
        with gr.Column():
            word_input = gr.File(label="📄 Upload Report Word Doc", file_types=[".docx"])
            index_dropdown = gr.Dropdown(
                choices=available_indexes,
                value=available_indexes[0] if available_indexes else None,
                label="📚 Select Reference Knowledge Base (RAG Index)",
                interactive=True
            )

            # --- API Key Input (Optional if using HF Secrets) ---
            # Uncomment the next 3 lines if you want users to enter their own key
            # api_key_box = gr.Textbox(
            #     label="🔑 Google API Key (optional if using HF Secrets)",
            #     type="password"
            # )

            # If using HF Secrets, hide the key box
            api_key_box = gr.Textbox(value="", visible=False)

            run_btn = gr.Button("🚀 Run Evaluation", variant="primary")

        with gr.Column():
            output_file = gr.File(label="📥 Download Evaluation Results (.xlsx)")

    # Connect inputs → function → output
    run_btn.click(
        fn=run_evaluation,
        inputs=[word_input, index_dropdown], # Pass word_input
        outputs=output_file
    )

    gr.Markdown("""
    ---
    **How it works**:
    - Your PDF is split into semantic chunks.
    - Each chunk is evaluated against the selected knowledge base using **Google Gemini**.
    - Results include accuracy scores, evidence, and expert commentary.
    - Average runtime 9 minutes.
    - 
    """)

# Launch (for local testing)
if __name__ == "__main__":
    app = demo
    demo.launch(
        server_name="0.0.0.0",
        server_port=7860,
        inbrowser=False,      # important in containers
        show_api=False,       # optional
        share=False           # disable public link
    )
